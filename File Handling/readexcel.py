import openpyxl

wb=openpyxl.load_workbook("C:\\Users\\HP\\Desktop\\readexcel.xlsx")
print(wb.sheetnames)
print("active sheet is:"+wb.active.title)
#create object of any sheet on which you want to work
sh=wb['Sheet1']
print(sh.title)
print(sh['A3'].value)
print(sh['B2'].value)
c1=sh.cell(1,1) #first is row second is column
print(c1.value)

c2=sh.cell(column=1,row=3)
print(c2.value)
print(c2.row)
print(c2.column)
